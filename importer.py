import json
import sqlite3
from datetime import datetime, timedelta
from sqlite3 import Error
from time import localtime

import openpyxl as openpyxl
import regex


class LogEntry:
    offset = timedelta(hours=(int(localtime().tm_gmtoff / 60 / 60)))

    def __init__(self, entry_id, date, body, offset=offset):
        self.id = entry_id
        self.date = datetime.utcfromtimestamp(date / 1000) + offset
        self.body = body
        self.op_type = None
        self.sum = 0
        self.category = ''
        self.cat_name = ''
        self.cat_comment = ''


def create_connection(db_file):
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except Error as e:
        print(e)

    return None


def select_entries(conn, last_id):
    last_id_param = (last_id,)

    cur = conn.cursor()
    cur.execute("SELECT _id, date, body FROM sms WHERE _id > ? and thread_id = 117 order by _id desc", last_id_param)

    rows = cur.fetchall()

    entries = []
    for row in rows:
        entry = LogEntry(row[0], row[1], row[2])
        entries.append(entry)

    return entries


def get_entries(config):
    conn = create_connection(config["general"]["db_path"])
    with conn:
        entries = select_entries(conn, config["general"]["last_id"])

    conn.close()

    if len(entries) > 0:
        config["general"]["last_id"] = entries[0].id

    entries.reverse()
    return entries


def process_entries(config, entries):
    entries_to_process = []
    for entry in entries:
        process(entry, config)
        if entry.op_type is None:
            print(entry.op_type, entry.id, entry.date, entry.body)
        elif entry.op_type is 1:
            entries_to_process.append(entry)

    if len(entries_to_process) > 0:
        write_entries_to_file(config, entries_to_process)


def process(entry, config):
    for type_def in config["types"]:
        for pattern in type_def["pattern"]:
            if pattern in entry.body:
                entry.op_type = type_def["id"]
                if entry.op_type == 1:
                    process_buy_entry(entry, config["categories"])
                else:
                    print(str(entry.date) + ": " + entry.body)


def process_buy_entry(entry, categories):
    parts = regex.match(r'.*\.\s\p{L}+\s([\d\.]*)\s\p{L}+\.\s(.*)\.\s\w+\s([\d\.]*)\s\w+\.\s(.*)', entry.body)
    entry.sum = float(parts.captures(1)[0])
    entry.category = parts.captures(2)[0]

    if entry.category in categories:
        entry.cat_name = categories[entry.category]["name"]
        entry.cat_comment = categories[entry.category]["comment"]


def write_entries_to_file(config, entries):
    sheet_path = config["general"]["sheet_path"]
    max_row = config["general"]["sheet_row"]

    wb = openpyxl.load_workbook(sheet_path)
    sheet = wb.worksheets[config["general"]["sheet_num"]]

    for entry in entries:
        max_row += 1
        write_entry_to_file(sheet, max_row, entry)

    wb.save(sheet_path)
    config["general"]["sheet_row"] = max_row


def write_entry_to_file(sheet, max_row, entry):
    sheet.cell(max_row, 1).value = entry.date.date()
    sheet.cell(max_row, 2).value = -1 * entry.sum
    if entry.cat_name != '':
        sheet.cell(max_row, 3).value = entry.cat_name
        sheet.cell(max_row, 4).value = entry.cat_comment
    else:
        sheet.cell(max_row, 5).value = entry.category


def main(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    entries = get_entries(config)
    process_entries(config, entries)

    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

    print('Done')


if __name__ == '__main__':
    main('config.json')
